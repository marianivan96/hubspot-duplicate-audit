import requests
import pandas as pd
from difflib import SequenceMatcher
from jinja2 import Template
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import time


from config import HUBSPOT_TOKEN

CONFIG = {
    "token":           HUBSPOT_TOKEN,
    "batch_size":      100,         # contacts per API call (max 200)
    "fuzzy_threshold": 0.85,        # name similarity score (0-1) for fuzzy match
    "output_html":     "hubspot_duplicate_audit.html",
    "output_excel":    "hubspot_duplicate_audit.xlsx",
}

HEADERS = {
    "Authorization": f"Bearer {CONFIG['token']}",
    "Content-Type": "application/json",
}

PROPERTIES = ["email", "firstname", "lastname", "company", "phone", "createdate", "hs_object_id"]


def fetch_all_contacts():
    """Pull all contacts from HubSpot using pagination."""
    contacts = []
    url = "https://api.hubapi.com/crm/v3/objects/contacts"  # works for both EU and US

    print("  Fetching contacts from HubSpot...")

    # Quick test call first
    test = requests.get(
        "https://api.hubapi.com/crm/v3/objects/contacts",
        headers=HEADERS,
        params={"limit": 2, "properties": "email,firstname,lastname"}
    )
    print(f"    → Test call status: {test.status_code}")
    if test.status_code == 200:
        test_data = test.json()
        print(f"    → Test total available: {test_data.get('total', 'N/A')}")
        print(f"    → Test results count: {len(test_data.get('results', []))}")
        if test_data.get("results"):
            print(f"    → Sample record keys: {list(test_data['results'][0].keys())}")
            print(f"    → Sample properties: {test_data['results'][0].get('properties', {})}")
    else:
        print(f"    → Error body: {test.text[:300]}")
        return []

    after = None
    page = 0

    while True:
        params = {
            "limit": CONFIG["batch_size"],
            "properties": ",".join(PROPERTIES),
        }
        if after:
            params["after"] = after

        response = requests.get(url, headers=HEADERS, params=params)

        if response.status_code != 200:
            print(f"  ✗ API error: {response.status_code} — {response.text}")
            break

        data = response.json()
        results = data.get("results", [])

        if page == 0:
            print(f"    → First page: {len(results)} contacts, keys in first record: {list(results[0].keys()) if results else 'EMPTY'}")

        if not results:
            break

        contacts.extend(results)
        page += 1

        if page % 10 == 0:
            print(f"    → {len(contacts)} contacts fetched so far...")

        # Get next page cursor
        paging = data.get("paging", {})
        next_page = paging.get("next", {})
        after = next_page.get("after") if next_page else None

        if not after:
            break

        time.sleep(0.05)

    print(f"  ✓ Total contacts fetched: {len(contacts)}")
    return contacts


def parse_contacts(raw):
    """Flatten API results into a clean DataFrame."""
    rows = []
    for c in raw:
        p = c.get("properties", {})
        # Handle both dict values and direct string values
        def get_prop(key):
            val = p.get(key)
            if val is None:
                return ""
            if isinstance(val, dict):
                return str(val.get("value", "")).strip()
            return str(val).strip()

        firstname = get_prop("firstname")
        lastname  = get_prop("lastname")
        fullname  = f"{firstname} {lastname}".strip()
        rows.append({
            "id":        str(c.get("id", "")),
            "email":     get_prop("email").lower(),
            "firstname": firstname,
            "lastname":  lastname,
            "fullname":  fullname,
            "company":   get_prop("company"),
            "phone":     get_prop("phone"),
            "createdate": get_prop("createdate")[:10] if get_prop("createdate") else "",
            "url":       f"https://app.hubspot.com/contacts/26134448/record/0-1/{c.get('id', '')}",
        })
    df = pd.DataFrame(rows)
    # Ensure all expected columns exist
    for col in ["id", "email", "firstname", "lastname", "fullname", "company", "phone", "createdate", "url"]:
        if col not in df.columns:
            df[col] = ""
    return df


def find_exact_email_dupes(df):
    """Type 1: Same email address on multiple records."""
    dupes = []
    email_df = df[df["email"] != ""].copy()
    grouped = email_df.groupby("email")
    for email, group in grouped:
        if len(group) > 1:
            records = group.to_dict("records")
            dupes.append({
                "type":    "Exact Email Match",
                "reason":  f"Email '{email}' appears {len(group)} times",
                "records": records,
                "count":   len(group),
            })
    return dupes

def find_same_name_dupes(df):
    """Type 2: Same full name, different email."""
    dupes = []
    named_df = df[(df["fullname"] != "") & (df["fullname"] != " ")].copy()
    grouped = named_df.groupby("fullname")
    for name, group in grouped:
        if len(name.strip()) < 4:
            continue
        if len(group) > 1:
            emails = group["email"].unique()
            if len(emails) > 1:
                records = group.to_dict("records")
                dupes.append({
                    "type":    "Same Name, Different Email",
                    "reason":  f"'{name}' has {len(group)} records with different emails",
                    "records": records,
                    "count":   len(group),
                })
    return dupes

def find_lastname_company_dupes(df):
    """Type 3: Same last name + company (catches job changers or entry errors)."""
    dupes = []
    sub = df[
        (df["lastname"] != "") &
        (df["company"] != "") &
        (df["lastname"].str.len() > 2)
    ].copy()
    sub["key"] = sub["lastname"].str.lower() + "|" + sub["company"].str.lower()
    grouped = sub.groupby("key")
    for key, group in grouped:
        if len(group) > 1:
            emails = group["email"].unique()
            if len(emails) > 1:
                lastname, company = key.split("|", 1)
                records = group.to_dict("records")
                dupes.append({
                    "type":    "Same Last Name + Company",
                    "reason":  f"Last name '{lastname}' at '{company}' has {len(group)} records",
                    "records": records,
                    "count":   len(group),
                })
    return dupes

def find_no_email_contacts(df):
    """Flag contacts with no email at all."""
    no_email = df[df["email"] == ""].copy()
    return no_email.to_dict("records")


def build_excel(df, all_dupes, no_email):
    """Generate Excel with 3 sheets: Summary, All Duplicates, No Email."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = "Summary"
    header_fill = PatternFill("solid", fgColor="1a6b4a")
    header_font = Font(bold=True, color="FFFFFF")

    summary_data = [
        ["Metric", "Count"],
        ["Total contacts", len(df)],
        ["Exact email duplicates (groups)", len([d for d in all_dupes if d["type"] == "Exact Email Match"])],
        ["Same name, diff email (groups)",  len([d for d in all_dupes if d["type"] == "Same Name, Different Email"])],
        ["Same last name + company (groups)", len([d for d in all_dupes if d["type"] == "Same Last Name + Company"])],
        ["Total duplicate groups",           len(all_dupes)],
        ["Contacts with no email",           len(no_email)],
        ["Report generated",                 datetime.now().strftime("%d %b %Y %H:%M")],
    ]
    for i, row in enumerate(summary_data, 1):
        for j, val in enumerate(row, 1):
            cell = ws1.cell(row=i, column=j, value=val)
            if i == 1:
                cell.fill = header_fill
                cell.font = header_font
    ws1.column_dimensions["A"].width = 35
    ws1.column_dimensions["B"].width = 15

    # ── Sheet 2: All Duplicates ──
    ws2 = wb.create_sheet("Duplicates")
    cols = ["Duplicate Type", "Reason", "ID", "Full Name", "Email", "Company", "Phone", "Created", "HubSpot URL"]
    for j, col in enumerate(cols, 1):
        cell = ws2.cell(row=1, column=j, value=col)
        cell.fill = header_fill
        cell.font = header_font

    row_idx = 2
    fill_colors = {"Exact Email Match": "FFE8E8", "Same Name, Different Email": "FFF3CD", "Same Last Name + Company": "E8F4FD"}
    for group in all_dupes:
        fill = PatternFill("solid", fgColor=fill_colors.get(group["type"], "FFFFFF"))
        for rec in group["records"]:
            vals = [
                group["type"], group["reason"],
                rec["id"], rec["fullname"], rec["email"],
                rec["company"], rec["phone"], rec["createdate"], rec["url"]
            ]
            for j, val in enumerate(vals, 1):
                cell = ws2.cell(row=row_idx, column=j, value=val)
                cell.fill = fill
            row_idx += 1
        row_idx += 1  # blank row between groups

    for col in ["A","B","C","D","E","F","G","H","I"]:
        ws2.column_dimensions[col].width = 28

    # ── Sheet 3: No Email ──
    ws3 = wb.create_sheet("No Email")
    cols3 = ["ID", "Full Name", "Company", "Phone", "Created", "HubSpot URL"]
    for j, col in enumerate(cols3, 1):
        cell = ws3.cell(row=1, column=j, value=col)
        cell.fill = header_fill
        cell.font = header_font
    for i, rec in enumerate(no_email, 2):
        vals = [rec["id"], rec["fullname"], rec["company"], rec["phone"], rec["createdate"], rec["url"]]
        for j, val in enumerate(vals, 1):
            ws3.cell(row=i, column=j, value=val)
    for col in ["A","B","C","D","E","F"]:
        ws3.column_dimensions[col].width = 28

    wb.save(CONFIG["output_excel"])
    print(f"  ✓ Excel saved: {CONFIG['output_excel']}")


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>HubSpot Duplicate Audit</title>
<style>
  :root {
    --ink: #0f1117; --paper: #f5f0eb; --green: #1a6b4a;
    --yellow: #b45309; --blue: #1d4ed8; --red: #b91c1c;
    --card: #ffffff; --border: #e5e7eb; --muted: #6b7280;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
         background: var(--paper); color: var(--ink); padding: 32px; }
  h1 { font-size: 28px; font-weight: 700; margin-bottom: 4px; }
  .subtitle { color: var(--muted); font-size: 14px; margin-bottom: 32px; }
  .stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(160px,1fr));
           gap: 16px; margin-bottom: 32px; }
  .stat-card { background: var(--card); border: 1px solid var(--border);
               border-radius: 8px; padding: 20px; }
  .stat-num { font-size: 32px; font-weight: 700; color: var(--green); }
  .stat-label { font-size: 12px; color: var(--muted); margin-top: 4px; }
  .section { margin-bottom: 40px; }
  .section-title { font-size: 18px; font-weight: 600; margin-bottom: 16px;
                   padding-bottom: 8px; border-bottom: 2px solid var(--border); }
  .dupe-group { background: var(--card); border: 1px solid var(--border);
                border-radius: 8px; margin-bottom: 16px; overflow: hidden; }
  .dupe-header { padding: 12px 16px; font-size: 13px; font-weight: 600;
                 display: flex; justify-content: space-between; }
  .type-email  { background: #fef2f2; color: var(--red); }
  .type-name   { background: #fffbeb; color: var(--yellow); }
  .type-company { background: #eff6ff; color: var(--blue); }
  table { width: 100%; border-collapse: collapse; font-size: 13px; }
  th { background: #f9fafb; padding: 8px 12px; text-align: left;
       font-weight: 600; color: var(--muted); border-bottom: 1px solid var(--border); }
  td { padding: 8px 12px; border-bottom: 1px solid #f3f4f6; }
  a { color: var(--green); text-decoration: none; }
  a:hover { text-decoration: underline; }
  .badge { display: inline-block; padding: 2px 8px; border-radius: 3px;
           font-size: 11px; font-weight: 600; }
  .badge-email   { background: #fef2f2; color: var(--red); }
  .badge-name    { background: #fffbeb; color: var(--yellow); }
  .badge-company { background: #eff6ff; color: var(--blue); }
  .no-email-table td, .no-email-table th { font-size: 12px; }
  footer { text-align: center; color: var(--muted); font-size: 12px; margin-top: 40px; }
</style>
</head>
<body>
<h1>🔍 HubSpot Duplicate Audit</h1>
<p class="subtitle">Generated {{ generated_at }} &nbsp;·&nbsp; {{ total_contacts }} total contacts analysed</p>

<div class="stats">
  <div class="stat-card">
    <div class="stat-num">{{ total_contacts }}</div>
    <div class="stat-label">Total Contacts</div>
  </div>
  <div class="stat-card">
    <div class="stat-num" style="color:var(--red)">{{ exact_email_count }}</div>
    <div class="stat-label">Exact Email Duplicate Groups</div>
  </div>
  <div class="stat-card">
    <div class="stat-num" style="color:var(--yellow)">{{ same_name_count }}</div>
    <div class="stat-label">Same Name, Diff Email Groups</div>
  </div>
  <div class="stat-card">
    <div class="stat-num" style="color:var(--blue)">{{ same_company_count }}</div>
    <div class="stat-label">Same Last Name + Company Groups</div>
  </div>
  <div class="stat-card">
    <div class="stat-num" style="color:var(--muted)">{{ no_email_count }}</div>
    <div class="stat-label">Contacts With No Email</div>
  </div>
</div>

{% for type_label, type_class, type_dupes in sections %}
{% if type_dupes %}
<div class="section">
  <div class="section-title">{{ type_label }} ({{ type_dupes|length }} groups)</div>
  {% for group in type_dupes %}
  <div class="dupe-group">
    <div class="dupe-header {{ type_class }}">
      <span>{{ group.reason }}</span>
      <span>{{ group.count }} records</span>
    </div>
    <table>
      <thead>
        <tr>
          <th>Name</th><th>Email</th><th>Company</th><th>Phone</th><th>Created</th><th>Link</th>
        </tr>
      </thead>
      <tbody>
        {% for rec in group.records %}
        <tr>
          <td>{{ rec.fullname or '—' }}</td>
          <td>{{ rec.email or '—' }}</td>
          <td>{{ rec.company or '—' }}</td>
          <td>{{ rec.phone or '—' }}</td>
          <td>{{ rec.createdate or '—' }}</td>
          <td><a href="{{ rec.url }}" target="_blank">View in HubSpot ↗</a></td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
  {% endfor %}
</div>
{% endif %}
{% endfor %}

{% if no_email %}
<div class="section">
  <div class="section-title">Contacts With No Email ({{ no_email|length }})</div>
  <div class="dupe-group">
    <table class="no-email-table">
      <thead><tr><th>Name</th><th>Company</th><th>Phone</th><th>Created</th><th>Link</th></tr></thead>
      <tbody>
        {% for rec in no_email[:100] %}
        <tr>
          <td>{{ rec.fullname or '—' }}</td>
          <td>{{ rec.company or '—' }}</td>
          <td>{{ rec.phone or '—' }}</td>
          <td>{{ rec.createdate or '—' }}</td>
          <td><a href="{{ rec.url }}" target="_blank">View ↗</a></td>
        </tr>
        {% endfor %}
        {% if no_email|length > 100 %}
        <tr><td colspan="5" style="color:var(--muted); padding:12px; text-align:center">
          ... and {{ no_email|length - 100 }} more. See Excel for full list.
        </td></tr>
        {% endif %}
      </tbody>
    </table>
  </div>
</div>
{% endif %}

<footer>hubspot_duplicate_audit.py &nbsp;·&nbsp; Data source: HubSpot CRM</footer>
</body>
</html>"""

def build_html(df, all_dupes, no_email):
    exact  = [d for d in all_dupes if d["type"] == "Exact Email Match"]
    names  = [d for d in all_dupes if d["type"] == "Same Name, Different Email"]
    company= [d for d in all_dupes if d["type"] == "Same Last Name + Company"]

    sections = [
        ("Exact Email Duplicates",         "type-email",   exact),
        ("Same Name, Different Email",      "type-name",    names),
        ("Same Last Name + Company",        "type-company", company),
    ]

    tmpl = Template(HTML_TEMPLATE)
    html = tmpl.render(
        generated_at=datetime.now().strftime("%d %b %Y, %H:%M"),
        total_contacts=len(df),
        exact_email_count=len(exact),
        same_name_count=len(names),
        same_company_count=len(company),
        no_email_count=len(no_email),
        sections=sections,
        no_email=no_email,
    )
    with open(CONFIG["output_html"], "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  ✓ HTML saved: {CONFIG['output_html']}")


def main():
    print("━" * 50)
    print("  HubSpot Duplicate Contact Audit")
    print("━" * 50)

    raw = fetch_all_contacts()
    df  = parse_contacts(raw)

    print("\n  Running duplicate checks...")
    exact_dupes   = find_exact_email_dupes(df)
    name_dupes    = find_same_name_dupes(df)
    company_dupes = find_lastname_company_dupes(df)
    no_email      = find_no_email_contacts(df)

    all_dupes = exact_dupes + name_dupes + company_dupes

    print(f"  → Exact email duplicates:     {len(exact_dupes)} groups")
    print(f"  → Same name, diff email:      {len(name_dupes)} groups")
    print(f"  → Same last name + company:   {len(company_dupes)} groups")
    print(f"  → Contacts with no email:     {len(no_email)}")

    print("\n  Building reports...")
    build_html(df, all_dupes, no_email)
    build_excel(df, all_dupes, no_email)

    print("\n" + "━" * 50)
    print("  ✅ Audit complete!")
    print(f"  Open {CONFIG['output_html']} in your browser")
    print("━" * 50)

if __name__ == "__main__":
    main()